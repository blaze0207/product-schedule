import pandas as pd
import glob
import os
import re
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reality_analyzer import RealityLogAnalyzer

def export_clear_plan():
    base_dir = os.getcwd()
    analyzer = RealityLogAnalyzer(base_dir)
    
    # 1. 尋找原始產銷表
    files = [f for f in glob.glob(os.path.join(base_dir, "*產銷*.xlsx")) if not os.path.basename(f).startswith('~$')]
    if not files:
        print("❌ 找不到包含 '產銷' 字樣的 Excel 檔案")
        return
    
    src_path = sorted(files, key=os.path.getmtime, reverse=True)[0]
    file_name = os.path.basename(src_path)
    
    # 移除現有的 _clear 後綴(如果有的話)再重新附加，確保檔名乾淨
    base_name, ext = os.path.splitext(file_name)
    clean_base = base_name.replace("_clear", "")
    output_path = os.path.join(base_dir, f"{clean_base}_clear{ext}")

    print(f"📂 正在美化處理原始檔案: {file_name}")

    # 2. 讀取並執行清洗邏輯
    xl = pd.ExcelFile(src_path)
    df = xl.parse(xl.sheet_names[0])
    
    # 補全機台名稱 (ffill)
    df.iloc[:, 0] = df.iloc[:, 0].ffill()
    
    # 2.1 執行機台過濾 (剔除雜訊)
    def is_valid_machine(m):
        m = str(m).upper().strip()
        if not m or m == 'NAN': return False
        if any(x in m for x in ['V', 'S2', '庫存', '待排', '機台', 'M4', '(CW)']):
            return False
        if re.search(r'M0[1-8]|S0[1-2]', m):
            return False
        return True

    machine_col = df.columns[0]
    df = df[df[machine_col].apply(is_valid_machine)].copy()

    # 針對批號欄位 (Index 1) 執行標準化清洗
    df['清洗後批號'] = df.iloc[:, 1].apply(lambda x: analyzer.standardize_id(x))
    
    # 針對 POY 批號 (Index 8) 補全
    if df.shape[1] > 8:
        df.iloc[:, 8] = df.iloc[:, 8].ffill()

    # 3. 輸出初步 Excel
    df.to_excel(output_path, index=False, engine='openpyxl')

    # 4. 使用 openpyxl 進行美化
    wb = load_workbook(output_path)
    ws = wb.active
    ws.title = "精簡版排產表"

    # 定義樣式
    header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid') # 深藍
    header_font = Font(color='FFFFFF', bold=True, size=12, name='微軟正黑體')
    data_font = Font(size=11, name='微軟正黑體')
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=False)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=False)
    
    thin_side = Side(border_style="thin", color="D9D9D9")
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    # 套用表頭樣式
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border

    # 套用資料樣式與自動調整欄寬
    for i, col in enumerate(ws.columns, 1):
        max_length = 0
        column = col[0].column_letter # 取得欄位字母 (A, B, C...)
        
        for cell in col:
            cell.border = border
            cell.font = data_font
            # 針對機台、日期、數量欄位置中，其餘靠左
            if i in [1, 4, 6, 7, 8, 11]: 
                cell.alignment = center_align
            else:
                cell.alignment = left_align
            
            # 計算最大寬度
            try:
                if cell.value:
                    val_len = len(str(cell.value).encode('utf-8')) # 考慮中文字長度
                    if val_len > max_length: max_length = val_len
            except: pass
        
        # 設定欄寬 (略大於文字寬度)
        adjusted_width = min(max_length + 2, 40) # 最高不超過 40
        ws.column_dimensions[column].width = adjusted_width

    # 凍結首行
    ws.freeze_panes = "A2"

    wb.save(output_path)
    print(f"✨ 旗艦版美化完成！已儲存至: {os.path.basename(output_path)}")
    return os.path.basename(output_path)

if __name__ == "__main__":
    export_clear_plan()
